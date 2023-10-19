import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings("ignore")
input_file = './out/out.xlsx'
doc=pd.read_excel(input_file)
# print(doc)

####################################
#lowest peak intensity
filtered_values = doc.iloc[:, 12].apply(pd.to_numeric, errors='coerce').dropna()
filtered_values = filtered_values[filtered_values != 0]
min_value = np.min(filtered_values)
point=min_value
# print(point)
# ####################################
df1=pd.DataFrame(doc.iloc[:,12].dropna().iloc[1:]) #50*1
# print(df1)
data_org=pd.DataFrame(doc.iloc[:,25:]) #3195*50
# print(data_org)
data = pd.DataFrame()

#dark 위치 찾기
count_zero = (df1 == 0).sum()
# print("0의 개수:", count_zero)
dark=data_org.iloc[:,count_zero]
# print(dark)
# ################################
for col in doc.columns[25:]:
    # print(doc[col])

    column_data= doc[col]


    # column_data = df.iloc[2:,col]
    # print(column_data)
    max_value = column_data.max()


    if max_value > point:
        data[col] = column_data
# max 최대값 넘는지 안넘는지
# print(data)
# ###########################################
counts = data.iloc[1].value_counts()
counts = counts.sort_index(ascending=False)
# print(counts)
counts_all=np.cumsum(counts)
df=pd.DataFrame()
wave=pd.DataFrame(doc.iloc[2:,24])
df['Wavelength']=wave
# print(wave)
#
for column in data.columns:
    if data[column].iloc[1]==counts.index[0]:
        # print(data[column])
        # print(dark.iloc[2:,0])
        datt=data[column].iloc[2:]-dark.iloc[2:,0]
        df[data[column].iloc[0]]=datt

    elif data[column].iloc[1]==counts.index[1]:
        # print(data.iloc[:,counts_all.iloc[0]])
        # print(df.iloc[:,counts_all.iloc[0]-1])
        datt_21=data[column].iloc[2:]-df.iloc[:,counts_all.iloc[0]-1]
        # print(datt_21)
        mean_values = datt_21[:50].mean()
        # print(mean_values)
        datt_22=data[column].iloc[2:]-mean_values
        # print(datt_22)
        df[data[column].iloc[0]]=datt_22
    for n in range(2,len(counts)):
        # print(n)
        if data[column].iloc[1]==counts.index[n]:
            # print(data.iloc[:,counts_all.iloc[n-1]-1])
            # print(data[column].iloc[2:])
            # print(counts_all.iloc[n - 1]-1)
            # print(df.iloc[:,counts_all.iloc[n - 1]-1])
            datt_31 = data[column].iloc[2:] - df.iloc[:, counts_all.iloc[n-1] - 1]
            # print(datt_31)
            mean_values = datt_31[:50].mean()
            # print(mean_values)
            datt_32 = data[column].iloc[2:] - mean_values
            # print(datt_32)
            df[data[column].iloc[0]] = datt_32
#
df.iloc[:, :] = df.iloc[:, :].applymap(lambda x: 0.01 if x <= 0 else x)
# print(df)
wave=df.iloc[:,0]
data=df.iloc[:,1:]
# print(data)
numwave=len(wave)
photonE=1240/wave
# print(photonE)
photondf=pd.DataFrame(photonE).reset_index()
meanlist=[]
for column in data.columns:
    # print(data[column].iloc[num-1])
    list=[]
    for a in range(numwave-2):
        # print(a)
        new_data=(photonE.iloc[a]-photonE.iloc[a+2])*0.5*(data[column].iloc[a]+data[column].iloc[a+2])
        list.append(new_data)
    res = [x * y for (x, y) in zip(photonE, list)]
    # print(len(res))
    # print(sum(res))
    sumlist=sum(list)
    a=sum(res)/sumlist
    # print(a)
    meanlist.insert(len(meanlist),a)
    list.insert(0,a)
    list.insert(1,0)
    # print(sumlist)
    # print(list)
    df2 = pd.DataFrame(list)
    # print(df2)
    photondf=pd.concat([photondf,df2],axis=1)
    # df.columns=new_header
photondf=photondf.iloc[:,1:]
photondf.columns=df.columns
# print(photondf)
# print(meanlist)
mean=pd.DataFrame(meanlist)
print(mean)
avg =pd.DataFrame([1240/x for x in meanlist])
# print(avg)
res=pd.concat([avg,mean],axis=1)
# print(res)
workbook = openpyxl.load_workbook(input_file)
worksheet = workbook['Sheet1']
for r_idx, row in enumerate(res.values, int(count_zero)+7):
    for c_idx, value in enumerate(row, 14):
        cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
new_sheet = workbook.create_sheet('Sheet2')
worksheet = workbook['Sheet2']
for row in dataframe_to_rows(photondf, index=False, header=True):
    worksheet.append(row)

workbook.save('./out/out_out.xlsx')
workbook.close()